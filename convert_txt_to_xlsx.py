import pandas as pd
import glob
import os
import chardet
import csv
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def detect_encoding(file_path):
    with open(file_path, 'rb') as f:
        result = chardet.detect(f.read())
    return result['encoding']

def detect_delimiter(file_path, encoding, default_delimiter=','):
    with open(file_path, 'r', encoding=encoding) as f:
        sample = f.read(1024)
        sniffer = csv.Sniffer()
        try:
            delimiter = sniffer.sniff(sample).delimiter
            # 如果检测到的分隔符是空格，可能需要手动确认
            if delimiter.isspace():
                print(f"检测到 {file_path} 使用空格作为分隔符")
        except csv.Error:
            print(f"无法检测到 {file_path} 的分隔符，使用默认分隔符 '{default_delimiter}'")
            delimiter = default_delimiter
    return delimiter

def add_borders_to_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    wb.save(file_path)

def convert_txt_to_xlsx():
    # 获取当前目录下所有的txt文件
    txt_files = glob.glob("*.txt")
    
    for txt_file in txt_files:
        # 检测文件的字符集
        encoding = detect_encoding(txt_file)
        
        # 检测文件的分隔符
        delimiter = detect_delimiter(txt_file, encoding)
        
        # 生成xlsx文件名（将.txt替换为.xlsx）
        xlsx_file = os.path.splitext(txt_file)[0] + '.xlsx'
        
        # 创建一个 Excel writer 对象
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            # 分块读取txt文件
            startrow = 0
            for chunk in pd.read_csv(txt_file, sep=delimiter, encoding=encoding, chunksize=10000):
                # 将每个块写入 Excel 文件
                chunk.to_excel(writer, index=False, startrow=startrow, header=startrow == 0)
                startrow += len(chunk)
        
        # 为所有单元格添加边框
        add_borders_to_excel(xlsx_file)
        
        print(f"已将 {txt_file} 转换为 {xlsx_file}")

if __name__ == "__main__":
    convert_txt_to_xlsx() 